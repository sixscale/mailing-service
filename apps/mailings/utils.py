import logging
from pathlib import Path
from typing import Generator

from openpyxl import load_workbook

from .exceptions import FileFormatError, RequiredColumnError

logger = logging.getLogger(__name__)


def read_xlsx_file(
    file_path: Path,
    required_columns: set[str],
) -> Generator[dict[str, str], None, None]:
    if not file_path.exists():
        raise FileFormatError(f"Файл не найден: {file_path}")

    if not file_path.suffix.lower() in (".xlsx", ".xls"):
        raise FileFormatError(f"Невалидный формат: {file_path.suffix}")

    try:
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    except Exception as e:
        raise FileFormatError(f"Не удалось прочитать файл: {e}")

    if not workbook.sheetnames:
        raise FileFormatError("В файле нет таблицы")

    worksheet = workbook.active

    header_row = next(worksheet.iter_rows(values_only=True), None)
    if not header_row:
        raise FileFormatError("файл пустой")

    headers = [str(cell).strip() if cell else "" for cell in header_row]

    header_set = set(headers)
    missing_columns = required_columns - header_set
    if missing_columns:
        raise RequiredColumnError(
            f"пропущены обязательные колоонки: {', '.join(sorted(missing_columns))}"
        )

    column_indices = {
        col: idx for idx, col in enumerate(headers) if col in required_columns
    }

    logger.info(
        "Чтение XLSX файла",
        extra={
            "file": str(file_path),
            "sheet": worksheet.title,
            "columns": headers,
        },
    )

    row_number = 1
    for row in worksheet.iter_rows(values_only=True):
        row_number += 1

        if not any(row):
            continue

        row_data = {}
        for col_name, col_idx in column_indices.items():
            value = row[col_idx] if col_idx < len(row) else None
            row_data[col_name] = str(value) if value is not None else ""

        yield row_data

    workbook.close()
